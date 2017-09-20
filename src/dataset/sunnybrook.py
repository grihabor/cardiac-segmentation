import os
import itertools
import logging
import tempfile
import zipfile
import shutil

from openpyxl import load_workbook

from dataset.dicom_reader import read_dicom_image
from utils import save_as_png

DIR_CACHE = '/data/cache'

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def filter_dicom(files, limit=None):
    filtered = filter(lambda filename: filename.endswith('.dcm'), files)
    if limit is None:
        return filtered
    else:
        return itertools.islice(filtered, limit)


def clean_cache():
    root_path = DIR_CACHE

    for root, dirs, files in os.walk(root_path):
        for filename in files:
            path = os.path.join(root_path, filename)
            try:
                os.remove(path)
                logger.debug(f'Remove {path}')
            except FileNotFoundError as e:
                logger.warning(e)


def read_sunnybrook_images(path):
    clean_cache()

    with zipfile.ZipFile(path, 'r') as zip_file:
        for zip_path in filter_dicom(zip_file.namelist(), limit=5):

            with tempfile.NamedTemporaryFile('wb') as temp_dicom:
                with zip_file.open(zip_path, 'r') as zipped_dicom:
                    shutil.copyfileobj(zipped_dicom, temp_dicom.file)
                img = read_dicom_image(temp_dicom.name)

            filename = zip_path.rsplit('/', 1)[-1]
            name = filename.split('.')[0]
            img_path = os.path.join(DIR_CACHE, f'{name}.png')
            save_as_png(img, img_path)


def read_sunnybrook_mapping(path):
    wb = load_workbook(path)
    ws = wb.worksheets[0]
    header = ws.rows[0]
    for row in ws.iter_rows():
        print(row)


def create_sunnybrook_tfrecords():
    read_sunnybrook_mapping()

